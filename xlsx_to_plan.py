#!/usr/bin/env python3
"""Konvertiert den Fitness-Masterplan (xlsx) in plan.json für die Trainings-App.

Aufruf:  python3 xlsx_to_plan.py "<pfad/zur/masterplan.xlsx>" [ausgabe.json]
Benötigt: pip install openpyxl
"""
import sys, json, re, unicodedata
import openpyxl

DAY_ORDER = ["Di", "Do", "Sa"]


def norm(s):
    s = unicodedata.normalize("NFKD", (s or "").lower())
    return re.sub(r"[^a-z0-9]+", "", s)


def split_semicolons_outside_parens(text):
    parts, depth, cur = [], 0, ""
    for ch in text:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
        if ch == ";" and depth == 0:
            parts.append(cur.strip())
            cur = ""
        else:
            cur += ch
    if cur.strip():
        parts.append(cur.strip())
    return parts


def parse_warmup(text):
    items = []
    for part in split_semicolons_outside_parens(text):
        m = re.match(r"^(\d+)\s*s\s+(.*)$", part)
        seconds = int(m.group(1)) if m else 30
        rest = m.group(2).strip() if m else part
        per_side = "je seite" in rest.lower()
        name = re.sub(r"\s*je Seite\s*", " ", rest, flags=re.I).strip()
        note = ""
        pm = re.match(r"^(.*?)\s*\((.*)\)\s*$", name)
        if pm:
            name, note = pm.group(1).strip(), pm.group(2).strip()
        items.append({"name": name, "seconds": seconds, "perSide": per_side, "note": note})
    return items


def parse_exercise_line(line):
    """z.B. 'A1 Einarmiges KB-Rudern vorgebeugt 4x6-10/Seite' -> code, name, sets, reps, perSide"""
    m = re.match(r"^([A-D][12])\s+(.*)$", line.strip())
    if not m:
        return None
    code, rest = m.group(1), m.group(2)
    sm = re.search(r"(\d+)\s*x\s*([\d\-]+m?)(/Seite)?", rest)
    sets = int(sm.group(1)) if sm else None
    reps = sm.group(2) if sm else ""
    per_side = bool(sm and sm.group(3)) or "/Seite" in rest
    name = rest[: sm.start()].strip(" –-") if sm else rest
    detail = rest[sm.end():].strip(" –-") if sm else ""
    return {"code": code, "name": name, "sets": sets, "targetReps": reps,
            "perSide": per_side, "detail": detail, "raw": line.strip()}


def parse_resistance(text):
    out = []
    for line in (text or "").split("\n"):
        line = line.strip()
        if not line:
            continue
        if ":" in line:
            label, val = line.split(":", 1)
            out.append({"label": label.strip(), "value": val.strip()})
        else:
            out.append({"label": "", "value": line})
    return out


def equipment_for_session(blocks):
    """Leitet die Vorbereitungs-Übersicht aus den Resistance-Angaben ab."""
    kb, dumbbell, barbell, bands, other = [], [], [], [], []
    for b in blocks:
        for r in b.get("resistance", []):
            v, l = r["value"], r["label"]
            entry = f"{v} ({l})" if l else v
            vl = v.lower()
            if "lbs" in vl or "band" in vl or "spannung" in vl or "gefühl" in vl:
                bands.append(entry)
            elif re.search(r"\b2x\s*[\d\-, ]+kg", vl) or "kurzhantel" in vl:
                dumbbell.append(entry)
            elif "langhantel" in vl or l.lower() in ("rudern", "rdl") and "kg" in vl:
                barbell.append(entry)
            elif "kb" in vl or l.lower().startswith("kb") or "kettlebell" in vl or l.lower() in ("goblet", "swing", "carry", "split squat", "sl-rdl"):
                kb.append(entry)
            elif "eigengewicht" in vl:
                pass
            else:
                other.append(entry)
    eq = []
    if dumbbell:
        eq.append({"gerät": "Kurzhanteln", "items": sorted(set(dumbbell))})
    if barbell:
        eq.append({"gerät": "Langhantel", "items": sorted(set(barbell))})
    if kb:
        eq.append({"gerät": "Kettlebell", "items": sorted(set(kb))})
    if bands:
        eq.append({"gerät": "Bänder", "items": sorted(set(bands))})
    if other:
        eq.append({"gerät": "Sonstiges", "items": sorted(set(other))})
    return eq


def main():
    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else "plan.json"
    wb = openpyxl.load_workbook(src, data_only=True)

    # ---------- Übungstabelle ----------
    exercises = []
    for row in wb["Übungstabelle"].iter_rows(min_row=2, values_only=True):
        name, desc, video = (row + ("", "", ""))[:3]
        if not name or not (desc or "").strip():
            continue
        name = str(name).strip()
        clean = re.sub(r"^NEU( v2)?:\s*", "", name).strip()
        video = str(video or "").strip()
        vm = re.search(r"https?://\S+", video)
        video_url = vm.group(0) if vm else ""
        video_note = video.replace(video_url, "").strip(" ()") if video_url else ""
        exercises.append({"name": clean, "description": str(desc).strip(),
                          "video": video_url, "videoNote": video_note})

    def find_exercise(label):
        n = norm(label)
        aliases = {
            "kbrudern": "einarmigeskbrudernvorgebeugt", "bss": "bulgariansplitsquat",
            "slrdl": "singlelegrdl", "rdl": "romaniandeadliftlanghantelv2",
            "rudern": "langhantelrudernvorgebeugt", "swing": "kbswingangepasst",
            "goblet": "gobletsquat", "carry": "suitcasecarry",
            "curl": "bandbizepscurl", "pushdown": "bandtrizepspushdown",
            "chestpress": "bandchestpressstehendeinarmig",
            "floorpress": "kurzhantelfloorpressneutralergriff",
            "aussenrotation": "bandaussenrotationen", "facepull": "bandfacepull",
            "lateralraise": "bandlateralraisebis80", "legraises": "liegendelegraises",
            "bandrudern": "bandrow", "pallof": "pallofpress",
            "splitsquat": "bulgariansplitsquat",
            "liegestutz": "liegestutzexzentrischphysio",
        }
        best = None
        for ex in exercises:
            en = norm(ex["name"])
            if n and (n in en or en in n):
                if best is None or len(en) < len(norm(best["name"])):
                    best = ex
        if best:
            return best
        for k, v in aliases.items():
            if k in n or n in k:
                for ex in exercises:
                    if norm(ex["name"]).startswith(v[:12]):
                        return ex
        return None

    # ---------- Workout Plan ----------
    sessions, cur = [], None
    for row in wb["Workout Plan"].iter_rows(min_row=2, values_only=True):
        day, block, details, week, resistance, _ = [(c if c is not None else "") for c in (row + ("",) * 6)[:6]]
        day, block, details = str(day).strip(), str(block).strip(), str(details).strip()
        if day.startswith("Progression") or (day.startswith("W1") and not block):
            break
        if day in DAY_ORDER and block == "Warm-up":
            cur = {"day": day, "week": None, "warmup": parse_warmup(details),
                   "blocks": [], "finish": None, "optional": None}
            sessions.append(cur)
            continue
        if cur is None:
            continue
        if block.startswith("Block"):
            if week != "" and cur["week"] is None:
                cur["week"] = int(week)
            res = parse_resistance(str(resistance))
            exs = []
            for i, line in enumerate([l for l in details.split("\n") if re.match(r"^[A-D][12]\s", l.strip())]):
                ex = parse_exercise_line(line)
                if not ex:
                    continue
                ex["resistance"] = res[i]["value"] if i < len(res) else ""
                ex["resistanceLabel"] = res[i]["label"] if i < len(res) else ""
                match = find_exercise(res[i]["label"] if i < len(res) and res[i]["label"] else ex["name"]) or find_exercise(ex["name"])
                ex["description"] = match["description"] if match else ""
                ex["video"] = match["video"] if match else ""
                ex["videoNote"] = match["videoNote"] if match else ""
                exs.append(ex)
            pause = ""
            pm = re.search(r"Pause\s*~?([\w\-, ]+s)", details)
            if pm:
                pause = "Pause ~" + pm.group(1).strip()
            note_lines = [l.strip() for l in details.split("\n")
                          if l.strip() and not re.match(r"^[A-D][12]\s", l.strip())]
            cur["blocks"].append({"block": block, "exercises": exs, "pause": pause,
                                  "note": " ".join(note_lines), "resistance": res})
        elif block in ("Finish", "Optional"):
            key = "finish" if block == "Finish" else "optional"
            cur[key] = {"text": details, "resistance": str(resistance).strip()}

    for s in sessions:
        s["equipment"] = equipment_for_session(s["blocks"])

    # ---------- Progression ----------
    progression = {}
    grab = False
    for row in wb["Workout Plan"].iter_rows(min_row=2, values_only=True):
        c0 = str(row[0] or "").strip()
        if c0.startswith("Progression"):
            grab = True
            continue
        if grab and re.match(r"^W\d+$", c0):
            progression[c0] = str(row[2] or "").strip()

    plan = {
        "meta": {
            "title": re.sub(r"\.xlsx$", "", src.split("/")[-1]),
            "source": src.split("/")[-1],
            "generated": __import__("datetime").date.today().isoformat(),
            # Anker für die Datumsrechnung der App: Montag der PLAN-Woche 12.
            # Verschoben von 2026-07-13 auf 2026-06-29, damit der August-Zyklus
            # (Wochen 16-19) schon in der Woche ab 27.07.2026 startet statt am 10.08.
            "week12Monday": "2026-06-29",
            "trainingDays": DAY_ORDER,
            "weeks": sorted({s["week"] for s in sessions}),
            "rules": "Rechter Arm über 90°/Schulterhöhe NUR bei den drei vom Physio verordneten Übungen (exzentrischer Klimmzug, exzentrischer Liegestütz, Physio 1) · sonst nie über Schulterhöhe · keine Cross-Body-Belastung · alles schmerzgeführt · Schmerz-Ampel entscheidet: grün ≤3/10 und binnen 24 h zurück · gelb = Sätze halbieren · rot = Übung raus und Physio fragen.",
        },
        "progression": progression,
        "sessions": sessions,
        "exercises": exercises,
    }
    with open(dst, "w", encoding="utf-8") as f:
        json.dump(plan, f, ensure_ascii=False, indent=1)
    print(f"OK: {len(sessions)} Sessions, {len(exercises)} Übungen -> {dst}")


if __name__ == "__main__":
    main()
